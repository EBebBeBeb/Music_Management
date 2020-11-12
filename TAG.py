# -*- coding: utf-8 -*-
import mutagen
from mutagen.mp3 import MP3, MPEGInfo,HeaderNotFoundError
from mutagen.flac import FLAC, StreamInfo,error
from mutagen.wave import WAVE, WaveStreamInfo
from mutagen.id3 import ID3, APIC, error,TALB,TDRC,TIT2,TPE1,TCOM,ID3NoHeaderError
from mutagen import MutagenError, StreamInfo
import os
import GUI
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from Web import Get_img
#APIC - Album PIC
#TALB = Title Album
#TDRC = 트랙년도 
#TIT2 = 타이틀
#TPE1 = 아티스트네임 
#TCOM = 작곡가 
def TagChecker(DIR):
	try:
		id3=ID3(DIR)
	except ID3NoHeaderError:
		id3=ID3()
	except MutagenError:
		return
	TAG_LIST=['TIT2','TPE1','TALB','TDRC','APIC','TCOM']
	TAG_NAME_LIST=['제목','아티스트','앨범','트랙년도','앨범아트','작곡가']
	TAG_STATUS=['X','X','X','X','X','X']
	print("파일 위치")
	print(DIR)
	for i in TAG_LIST:
		print(TAG_NAME_LIST[TAG_LIST.index(i)])
		if id3.getall(i)==[]:
			print("TAG MISSED")
		else:
			TAG_STATUS[TAG_LIST.index(i)]='O'
			if i == 'APIC':
				print('있음')
			else:
				print(id3.getall(i)[0])
	print('\n')
	return TAG_STATUS

def LengthLoader(DIR,EXT):
	if EXT == ".mp3":
		try:
			File=MP3(DIR)
		except HeaderNotFoundError:
			return 0
		return File.info.length
	elif EXT ==".flac":
		try:
			File=FLAC(DIR)
		except mutagen.flac.error:
			return 0
		return File.info.length
	elif EXT ==".wav":
		File=WAVE(DIR)
		return File.info.length

def M_TagWrite(DIR,TAG,Text):
	try:
		id3=ID3(DIR)
	except ID3NoHeaderError:
		id3=ID3()
	except MutagenError:
		return

	APIC_Flag=0

	if TAG == 0:
		id3["TIT2"] = TIT2(enconding=3,text=Text)
	elif TAG ==1:
		id3["TPE1"] = TPE1(enconding=3,text=Text)
	elif TAG ==2:
		id3["TALB"] = TALB(enconding=3,text=Text)
	elif TAG ==3:
		id3["TDRC"] = TDRC(enconding=3,text=Text)
	elif TAG ==4:
		id3["APIC"] = APIC(enconding=3,mime='image/jpeg',type=3,desc=u'Cover',data=open(Get_img(Text),'rb').read())
		APIC_Flag=1
	elif TAG ==5:
		id3["TCOM"] = TCOM(enconding=3,text=Text)
	id3.save(DIR)
	if APIC_Flag==1:
		os.remove(test.jpg)
		
def XYL_Writer(wb,DIR,num):
	try:
		id3=ID3(DIR)
	except ID3NoHeaderError:
		id3=ID3()
	except MutagenError:
		return
	ws=wb.active
	data=[]
	ws.title="MUSIC"
	TAG_LIST=['TIT2','TPE1','TALB','TDRC','APIC','TCOM']
	print(DIR)
	data.append(DIR)
	for i in TAG_LIST:
		if id3.getall(i)==[]:
			data.append('없음')
		else:
			if i == 'APIC':
				data.append('있음')
			else:
				data.append(str(id3.getall(i)[0]))
	ws.append(data)

def XYL_INIT():
	wb=Workbook()
	ws=wb.active
	ws1=wb.create_sheet("MUSIC")
	ws.title="MUSIC"
	ws['A1']='상대경로'
	ws['B1']='제목'
	ws['C1']='아티스트'
	ws['D1']='앨범'
	ws['E1']='트랙년도'
	ws['F1']='앨범아트 유무'
	ws['G1']='작곡가'
	return wb


def XYL_Reader(ws,num):
	data=[]
	for column in range(1,8):
		column_letter = get_column_letter(column)
		data.append(str(ws[column_letter+str(num)].value))
	return data

def XYL_Apply(DIR):
	wb=load_workbook(DIR)
	ws=wb.active
	sheet = wb.worksheets[0]
	for i in range(2,sheet.max_row+1):
		data=XYL_Reader(ws,i)
		print(data[0])
		for j in range(1,7):
			if data[j] != '없음' and data[j] != '있음':
				M_TagWrite(data[0],j-1,data[j])

